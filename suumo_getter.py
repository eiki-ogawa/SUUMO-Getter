# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import chromedriver_binary
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


class SuumoGetter():


    def __init__(self):
        """
        ドライブの設定
        待機時間など
        """
        self.driver = webdriver.Chrome()
        self.wait = WebDriverWait(self.driver, 30)

    def click(self, elem):
        """
        安定したクリックを行う関数
        前後で全て読み込むまで待機する時間を設けている
        画面外でクリックできるようにexecute_scriptを用いる。

        Parameters
        ----------
        elem : web-element
            クリックする要素
        """
        self.wait.until(EC.presence_of_all_elements_located) # すべて表示されるまで待機
        self.driver.execute_script("arguments[0].click();", elem)
        self.wait.until(EC.presence_of_all_elements_located) # すべて表示されるまで待機

    def rental_search(self, line, station):
        """
        賃貸情報を検索する

        Parameters
        ----------
        line : str
            沿線（例：ＪＲ山手線）
        station : str
            駅名（例：秋葉原）
        """
        # 東京沿線でSUUMO検索
        self.driver.get("https://suumo.jp/chintai/tokyo/ensen/")   
        self.wait.until(EC.presence_of_all_elements_located) # すべて表示されるまで待機

        # 沿線の要素を探す
        line_elem = self.driver.find_element(by=By.XPATH, value='//*[contains(text(), "{}")]'.format(line))
        self.click(line_elem)
        # 駅名を探す
        station_name = self.driver.find_element(by=By.XPATH, value='//li/label/span[contains(text(), "{}")]'.format(station))
        # チェックボックスの要素を探す
        station_elem = station_name.find_element(by=By.XPATH, value="../../input")
        self.click(station_elem)
        
        # 検索
        search_elem = self.driver.find_elements(by=By.XPATH, value='//*[contains(text(), "この条件で検索する")]')
        for i in search_elem:
            # 失敗した場合、次の検索ボタンを押す
            try:
                self.click(i)
                num_elem = self.driver.find_element(by=By.XPATH, value="//div[@class='paginate_set-hit']")
                break
            except:
                continue

    def rental_get(self, limit):
        """
        賃貸情報を取得する

        Parameters
        ----------
        limit: int
            取得する情報の上限
        
        Returns
        -------
        rental_data : list
            賃貸情報のリスト
        """
        # 賃貸情報のリスト
        rental_data = []

        try:
            while True:

                # 賃貸情報のテーブル
                rental_table = self.driver.find_elements(by=By.XPATH, value="//table[@class='cassetteitem_other']")
                # テーブルから必要な行（tbody）のみを取得
                for table in rental_table:
                    rental_tbody = table.find_elements(by=By.XPATH, value="tbody")

                    # 取得した行（tbody）から必要な情報のみ取得
                    for tbody in rental_tbody:
                        try:
                            # 駅まで何分か（一番近い駅のみ）
                            minutes = tbody.find_element(by=By.XPATH, value="../../../div[@class='cassetteitem-detail']/div[2]/div/div[3]/ul/li[2]/div[1]").get_attribute("textContent")
                            minutes = minutes.split(" ")[1]

                            # 築年数
                            age = tbody.find_element(by=By.XPATH, value="../../../div[@class='cassetteitem-detail']/div[2]/div/div[3]/ul/li[3]/div[1]").get_attribute("textContent")[1:-1]
                            if age == "":
                                # 新築の場合
                                age = 0

                            # 階数
                            floor = tbody.find_element(by=By.XPATH, value="tr/td[3]").get_attribute("textContent")
                            floor = floor.replace("階","").replace("\t","").replace("\n","")
                            try:
                                floor = int(floor)
                            except:
                                # 1-2 階と書かれていた場合や、地下にある物件など
                                pass

                            # 賃料
                            rent = tbody.find_element(by=By.XPATH, value="tr/td[4]/ul/li[1]/span/span").get_attribute("textContent")
                            rent = rent.replace("万円","")

                            # 管理費
                            manage_fee = tbody.find_element(by=By.XPATH, value="tr/td[4]/ul/li[2]/span").get_attribute("textContent")
                            manage_fee = manage_fee.replace("円","").replace("-","0")

                            # 敷金
                            deposit = tbody.find_element(by=By.XPATH, value="tr/td[5]/ul/li[1]/span").get_attribute("textContent")
                            deposit = deposit.replace("万円", "").replace("-","0")

                            # 礼金
                            key_money = tbody.find_element(by=By.XPATH, value="tr/td[5]/ul/li[2]/span").get_attribute("textContent")
                            key_money = key_money.replace("万円", "").replace("-","0")

                            # 間取り
                            floor_plan = tbody.find_element(by=By.XPATH, value="tr/td[6]/ul/li[1]/span").get_attribute("textContent")

                            # 専有面積
                            area = tbody.find_element(by=By.XPATH, value="tr/td[6]/ul/li[2]/span").get_attribute("textContent")[:-2]
                            area = area.replace("m2", "")

                            # リンク
                            link =tbody.find_element(by=By.XPATH, value="tr/td[9]/a[contains(text(), '詳細を見る')]").get_attribute("href")

                            # 平米単価 =(家賃+管理費)/専有面積
                            area_price = round((float(rent)*10000 + int(manage_fee))/float(area))

                            # 1物件ごとに 1つのリストにまとめる
                            data = [area_price, minutes, int(age), floor, round(float(rent)*10000), int(manage_fee), round(float(deposit)*10000), round(float(key_money)*10000), float(area), floor_plan, link]
                            # 出力用リストに追加
                            rental_data.append(data)
                            print(data)

                            # 指定件で中断
                            if len(rental_data) == limit:
                                return True

                        # データ取得中にエラーが発生した場合
                        # エラーをプリントして飛ばす
                        except Exception as e:
                            print(e)
                            pass

                #「次へ」を押す
                try:
                    next_elem = self.driver.find_element(by=By.XPATH, value="//div[@class='pagination pagination_set-nav']/p/a[contains(text(), '次へ')]")
                    self.click(next_elem)
                # 無かったら抜け出す
                except:
                    break

        except:
            pass
        # ドライブの終了
        finally:
            try:
                self.driver.quit()
                self.driver = None
            except:
                pass
            # 賃貸情報のリストを返す
            return rental_data

    def export_excel(self, rental_data):
        """
        Excel に出力する

        Parameters
        ----------
        rental_data : list
        SUUMO から取得した賃貸情報のリスト
        """
        # 平米単価で昇順に並べ替え
        rental_data = sorted(rental_data, key=lambda x: x[0])

        # 見出しの追加
        dict_list = ["平米単価","駅まで","築年数","階数","賃料","管理費","敷金","礼金","専有面積","間取り","リンク"]
        rental_data.insert(0, dict_list)

        # Excel 操作
        wb = openpyxl.Workbook()
        ws = wb["Sheet"]

        for i in range(len(rental_data)):
            for j in range(len(dict_list)):
                # データを入力
                ws.cell(row=i+1, column=j+1, value=rental_data[i][j])

        # セル幅の設定
        for col in ws.columns:
            max_length = 7
            column = col[0].column_letter

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = round((max_length + 2) * 1.2)
            ws.column_dimensions[column].width = adjusted_width

        # フィルターを設定
        ws.auto_filter.ref = get_column_letter(1) + str(1) + ':' + get_column_letter(ws.max_column) + str(ws.max_row)
        # 1行目を固定
        ws.freeze_panes = "A2"

        # ハイパーリンクの設定
        for i in range(len(rental_data)):
            if i != 0:
                link_cell = ws.cell(row=i+1, column=11)
                link_cell.hyperlink = link_cell.value

        # 保存
        wb.save("rental.xlsx")
        print("保存が完了しました。終了します。")

    def main(self, line, station, limit):
        """
        メイン関数
        
        Parameters
        ----------
        line : str
            沿線（例：ＪＲ山手線）
        station : str
            駅名（例：秋葉原）
        limit : int
            取得する情報の上限
        """
        # 賃貸情報を検索する
        self.rental_search(line, station)
        # 賃貸情報を取得する
        rental_data = self.rental_get(limit)
        # Excel に出力する
        if rental_data != []:
            self.export_excel(rental_data)

if __name__ == "__main__":
    sg = SuumoGetter()
    sg.main("ＪＲ中央線","吉祥寺", 100)